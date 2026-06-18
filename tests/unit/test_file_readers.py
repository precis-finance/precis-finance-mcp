# SPDX-License-Identifier: Elastic-2.0
# Copyright (c) 2026 Sergio Naval Marimont
"""Tests for csv, parquet, xlsx file readers."""

from __future__ import annotations

import io

import pytest

from precis_mcp.ingestion.file_readers import (
    CsvReader,
    ParquetReader,
    XlsxReader,
    get_reader,
)
from precis_mcp.ingestion.file_readers.base import FileReadError, apply_column_map


# ---------------------------------------------------------------------------
# apply_column_map — shared helper
# ---------------------------------------------------------------------------


def test_apply_column_map_renames_keys():
    raw = {"Date": "2026-04-01", "Account": "1000", "Amount": "100.00"}
    out = apply_column_map(
        raw,
        {"Date": "posting_date", "Account": "account_code", "Amount": "amount"},
    )
    assert out == {
        "posting_date": "2026-04-01",
        "account_code": "1000",
        "amount": "100.00",
    }


def test_apply_column_map_identity_passes_through():
    raw = {"posting_date": "x", "amount": 10}
    assert apply_column_map(raw, "identity") == raw


def test_apply_column_map_drops_unmapped_source_fields():
    raw = {"Date": "x", "Junk": "y", "Amount": 1}
    out = apply_column_map(raw, {"Date": "posting_date", "Amount": "amount"})
    assert "junk" not in out
    assert "Junk" not in out
    assert "posting_date" in out and "amount" in out


def test_apply_column_map_rejects_bad_type():
    with pytest.raises(FileReadError):
        apply_column_map({"x": 1}, 12345)  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_get_reader_returns_instances():
    assert isinstance(get_reader("csv"), CsvReader)
    assert isinstance(get_reader("parquet"), ParquetReader)
    assert isinstance(get_reader("xlsx"), XlsxReader)


def test_get_reader_unknown_format_raises():
    with pytest.raises(KeyError, match="ghost"):
        get_reader("ghost")


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _csv_bytes() -> bytes:
    return (
        "Date,Account,Amount,DR_CR\n"
        "2026-04-01,1000,1000.00,D\n"
        "2026-04-01,4000,1000.00,C\n"
    ).encode("utf-8")


def test_csv_reads_with_header_and_column_map():
    reader = CsvReader()
    rows = list(
        reader.read(
            _csv_bytes(),
            format_config={"delimiter": ",", "has_header": True},
            column_map={
                "Date": "posting_date",
                "Account": "account_code",
                "Amount": "amount",
                "DR_CR": "debit_credit",
            },
        )
    )
    assert len(rows) == 2
    assert rows[0]["posting_date"] == "2026-04-01"
    assert rows[0]["account_code"] == "1000"
    assert rows[0]["debit_credit"] == "D"
    assert rows[1]["account_code"] == "4000"


def test_csv_null_marker_becomes_none():
    reader = CsvReader()
    body = b"a,b\n1,\\N\n"
    rows = list(
        reader.read(
            body,
            format_config={"null_marker": r"\N"},
            column_map="identity",
        )
    )
    assert rows == [{"a": "1", "b": None}]


def test_csv_identity_column_map():
    reader = CsvReader()
    rows = list(
        reader.read(
            b"posting_date,amount\n2026-04-01,100\n",
            format_config={},
            column_map="identity",
        )
    )
    assert rows == [{"posting_date": "2026-04-01", "amount": "100"}]


def test_csv_accepts_binary_io():
    reader = CsvReader()
    stream = io.BytesIO(_csv_bytes())
    rows = list(
        reader.read(stream, format_config={}, column_map="identity")
    )
    assert len(rows) == 2


def test_csv_custom_delimiter():
    reader = CsvReader()
    body = b"a;b\n1;2\n"
    rows = list(
        reader.read(body, format_config={"delimiter": ";"}, column_map="identity")
    )
    assert rows == [{"a": "1", "b": "2"}]


def test_csv_no_header_requires_columns():
    reader = CsvReader()
    with pytest.raises(FileReadError, match="columns"):
        list(reader.read(b"1,2,3\n", format_config={"has_header": False}, column_map="identity"))


def test_csv_no_header_explicit_columns():
    reader = CsvReader()
    rows = list(
        reader.read(
            b"1,2\n3,4\n",
            format_config={"has_header": False, "columns": ["a", "b"]},
            column_map="identity",
        )
    )
    assert rows == [{"a": "1", "b": "2"}, {"a": "3", "b": "4"}]


# ---------------------------------------------------------------------------
# Parquet
# ---------------------------------------------------------------------------


def _parquet_bytes() -> bytes:
    import pyarrow as pa
    import pyarrow.parquet as pq

    table = pa.table(
        {
            "Date": ["2026-04-01", "2026-04-01"],
            "Account": ["1000", "4000"],
            "Amount": [1000.0, 1000.0],
            "DR_CR": ["D", "C"],
        }
    )
    buf = io.BytesIO()
    pq.write_table(table, buf)
    return buf.getvalue()


def test_parquet_reads_and_maps():
    reader = ParquetReader()
    rows = list(
        reader.read(
            _parquet_bytes(),
            format_config={},
            column_map={
                "Date": "posting_date",
                "Account": "account_code",
                "Amount": "amount",
                "DR_CR": "debit_credit",
            },
        )
    )
    assert len(rows) == 2
    assert rows[0]["account_code"] == "1000"
    assert rows[0]["amount"] == 1000.0


def test_parquet_projection_via_columns():
    reader = ParquetReader()
    rows = list(
        reader.read(
            _parquet_bytes(),
            format_config={"columns": ["Date", "Amount"]},
            column_map={"Date": "posting_date", "Amount": "amount"},
        )
    )
    assert rows[0].keys() == {"posting_date", "amount"}


def test_parquet_malformed_bytes_raises():
    reader = ParquetReader()
    with pytest.raises(FileReadError, match="parquet parse"):
        list(reader.read(b"not parquet", format_config={}, column_map="identity"))


# ---------------------------------------------------------------------------
# xlsx
# ---------------------------------------------------------------------------


def _xlsx_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Account", "Amount", "DR_CR"])
    ws.append(["2026-04-01", "1000", 1000.0, "D"])
    ws.append(["2026-04-01", "4000", 1000.0, "C"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_xlsx_reads_first_sheet_with_header():
    reader = XlsxReader()
    rows = list(
        reader.read(
            _xlsx_bytes(),
            format_config={},
            column_map={
                "Date": "posting_date",
                "Account": "account_code",
                "Amount": "amount",
                "DR_CR": "debit_credit",
            },
        )
    )
    assert len(rows) == 2
    assert rows[0]["account_code"] == "1000"
    assert rows[1]["debit_credit"] == "C"


def test_xlsx_skip_rows():
    """A workbook with two preamble rows before the header."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["preamble", "info"])
    ws.append(["more", "stuff"])
    ws.append(["A", "B"])
    ws.append(["1", "2"])
    buf = io.BytesIO()
    wb.save(buf)
    rows = list(
        XlsxReader().read(
            buf.getvalue(),
            format_config={"skip_rows": 2},
            column_map="identity",
        )
    )
    assert rows == [{"A": "1", "B": "2"}]


def test_xlsx_named_sheet():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "primary"
    ws.append(["A"])
    ws.append([1])
    second = wb.create_sheet("secondary")
    second.append(["B"])
    second.append([2])
    buf = io.BytesIO()
    wb.save(buf)
    rows = list(
        XlsxReader().read(
            buf.getvalue(),
            format_config={"sheet": "secondary"},
            column_map="identity",
        )
    )
    assert rows == [{"B": 2}]


def test_xlsx_unknown_sheet_raises():
    reader = XlsxReader()
    with pytest.raises(FileReadError, match="sheet"):
        list(
            reader.read(
                _xlsx_bytes(),
                format_config={"sheet": "ghost"},
                column_map="identity",
            )
        )


def test_xlsx_malformed_bytes_raises():
    reader = XlsxReader()
    with pytest.raises(FileReadError, match="xlsx parse"):
        list(reader.read(b"not xlsx", format_config={}, column_map="identity"))


def test_xlsx_skips_trailing_blank_rows():
    """openpyxl emits None tuples for blank trailing rows; we must skip them."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, 2])
    # Forcing a trailing blank: openpyxl normally skips truly-empty rows, but
    # rows with explicit None are surfaced. We construct one defensively.
    ws.append([None, None])
    buf = io.BytesIO()
    wb.save(buf)
    rows = list(
        XlsxReader().read(buf.getvalue(), format_config={}, column_map="identity")
    )
    assert rows == [{"A": 1, "B": 2}]
