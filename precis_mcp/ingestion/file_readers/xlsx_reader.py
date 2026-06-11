# SPDX-License-Identifier: Elastic-2.0
# Copyright (c) 2026 Sergio Naval Marimont
"""xlsx reader (openpyxl-backed)."""

from __future__ import annotations

import io
from typing import Any, BinaryIO, Iterable, Union

from precis_mcp.ingestion.file_readers.base import (
    ColumnMap,
    FileReadError,
    apply_column_map,
)


class XlsxReader:
    """Single-workbook xlsx via openpyxl.

    Accepted `format_config` keys:
      - `sheet` — sheet name or index (default: first sheet)
      - `has_header` (default True). When False, `columns` must be supplied.
      - `skip_rows` (default 0) — rows to skip before the header / first data row
    """

    format_name = "xlsx"

    def read(
        self,
        file_data: Union[bytes, BinaryIO],
        *,
        format_config: dict[str, Any],
        column_map: ColumnMap,
    ) -> Iterable[dict[str, Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover — openpyxl is in requirements
            raise FileReadError(
                "xlsx reader requires openpyxl; install openpyxl"
            ) from exc

        if isinstance(file_data, (bytes, bytearray)):
            source: Any = io.BytesIO(file_data)
        else:
            source = file_data

        try:
            wb = load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            raise FileReadError(f"xlsx parse failure: {exc}") from exc

        sheet_ref = format_config.get("sheet")
        if sheet_ref is None:
            ws = wb.worksheets[0]
        elif isinstance(sheet_ref, int):
            ws = wb.worksheets[sheet_ref]
        else:
            try:
                ws = wb[sheet_ref]
            except KeyError as exc:
                raise FileReadError(f"xlsx sheet {sheet_ref!r} not found") from exc

        skip_rows = int(format_config.get("skip_rows", 0))
        has_header = format_config.get("has_header", True)

        rows_iter = ws.iter_rows(values_only=True)
        for _ in range(skip_rows):
            next(rows_iter, None)

        if has_header:
            header_row = next(rows_iter, None)
            if header_row is None:
                return []
            columns = [str(c) if c is not None else "" for c in header_row]
        else:
            columns = list(format_config.get("columns") or [])
            if not columns:
                raise FileReadError(
                    "xlsx reader: has_header=False requires 'columns' in format_config"
                )

        out: list[dict[str, Any]] = []
        for raw in rows_iter:
            if all(c is None for c in raw):
                continue  # Skip trailing blank rows.
            row_dict = dict(zip(columns, raw))
            out.append(apply_column_map(row_dict, column_map))
        return out
